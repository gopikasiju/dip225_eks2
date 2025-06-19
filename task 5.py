from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx', data_only=True)
ws = wb['Lapa_0']

total_sum = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    client = row[5]     # Column F
    quantity = row[11]  # Column L
    total = row[13]     # Column N

    if client == 'Korporatīvais' and isinstance(quantity, (int, float)) and 40 <= quantity <= 50 and isinstance(total, (int, float)):
        total_sum += total

print(int(total_sum))
