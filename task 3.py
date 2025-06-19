from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

count = sum(
    1 for row in ws.iter_rows(min_row=2, values_only=True)
    if isinstance(row[3], str) and 'Adulienas iela' in row[3] and row[4] in ('Valmiera', 'Saulkrasti')
)

print(count)
